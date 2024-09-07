import openpyxl # type: ignore


def getRowCount(file,sheetName):    #take count (lenght) of row
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_row)

def getRowCount(file,sheetName):    ##take count (lenght) of column
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,colnum): #read the data from exel based on row no and col no so we get aurate cell data
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(row=rownum,column=colnum).value

def writeData(file,sheetName,rownum,colnum,data):   #write the data from exel based on row no and col no so we get
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(row=rownum,column=colnum).value=data
    workbook.save(file)

